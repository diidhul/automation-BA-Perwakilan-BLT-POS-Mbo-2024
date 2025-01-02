import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os


# Fungsi untuk menggantikan placeholder dalam file Excel
def fill_excel_from_csv(csv_file, excel_template, output_file):
    try:
        # Membaca data dari file CSV dan memastikan kolom NIK Pengganti diproses sebagai string
        df = pd.read_csv(csv_file, dtype={"NIK Pengganti": str})

        # Memuat template file Excel
        workbook = load_workbook(excel_template)
        sheet = workbook.active  # Menggunakan sheet pertama

        # Pastikan semua merge range di template tetap ada
        for merge_range in sheet.merged_cells.ranges:
            sheet.merge_cells(str(merge_range))

        # Mengisi cell D16 dengan nama kelurahan pertama dari file CSV
        if not df.empty:  # Pastikan file CSV tidak kosong
            kelurahan_name = df.iloc[0][
                "Kelurahan"
            ]  # Ambil kelurahan dari baris pertama
            sheet["D16"].value = (
                f": {kelurahan_name}"  # Isi cell D16 dengan nama kelurahan
            )

        # Baris awal untuk data di template Excel (disesuaikan dengan header di baris ke-23)
        start_row = 24  # Data dimulai dari baris ke-24 (setelah header)
        data_row_count = len(df)

        # Iterasi setiap baris di CSV untuk mengisi data di kolom tertentu sesuai template
        for index, row in df.iterrows():
            # Menambahkan nomor urut di kolom A
            sheet.cell(row=start_row + index, column=1).value = index + 1

            # Kolom yang ingin diisi
            values = [
                (2, row.get("No_Urut_BAST", "")),  # Kolom B
                (3, row.get("Nama Penerima", "")),  # Kolom C
                (4, row.get("Nama Pengganti", "")),  # Kolom D
                (5, row.get("Nomor Cekpos", "")),  # Kolom E
                (6, f"'{row.get('NIK Pengganti', '')}"),  # Kolom F sebagai teks
                (7, row.get("Kelurahan", "")),  # Kolom G
                (8, row.get("Kelurahan", "")),  # Kolom H
            ]

            for col, value in values:
                cell = sheet.cell(row=start_row + index, column=col)
                cell.value = value

        # Menambahkan border, perataan tengah, dan font di bawah baris 42
        border = Border(
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        font = Font(name="Times New Roman", size=11)
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")

        # Mengaplikasikan styling ke baris di bawah 42
        for row in range(start_row, start_row + data_row_count):
            for col in range(
                1, 10
            ):  # Asumsi data berada di kolom 1-9 (termasuk kolom H dan I)
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.font = font

                # Untuk kolom Nama Penerima (C) dan Nama Pengganti (D) rata kiri
                if col == 3 or col == 4:
                    cell.alignment = alignment_left
                else:
                    cell.alignment = alignment_center

            # Mengatur tinggi baris menjadi 18.6 untuk setiap baris data
            sheet.row_dimensions[row].height = 18.60

        # Simpan file Excel yang sudah diisi
        workbook.save(output_file)
        return f"File Excel berhasil dibuat: {output_file}"

    except Exception as e:
        return f"Terjadi kesalahan: {e}"


# Fungsi utama untuk memproses semua file CSV di folder dan menyimpan output di folder yang sama
def process_all_csv_in_folder(input_folder, excel_template, output_folder):
    try:
        # Pastikan folder output utama ada
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Iterasi setiap file CSV di folder input
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".csv"):
                csv_file = os.path.join(input_folder, file_name)
                base_name = os.path.splitext(file_name)[0]  # Nama file tanpa ekstensi

                # Path output Excel
                output_file = os.path.join(output_folder, f"{base_name}.xlsx")

                # Jalankan pengisian template
                result = fill_excel_from_csv(csv_file, excel_template, output_file)
                print(result)

        print("\nSemua file BA Siap yaaa Dini Sayaang!!!")
        print("\nSkarang kamu jalanin 👉 python tambahinFooter.py")
        print("\n\n")

    except Exception as e:
        print(f"Terjadi kesalahan: {e}")


if __name__ == "__main__":
    # Input folder CSV, template Excel, dan folder output utama
    input_folder = os.path.join(os.getcwd(), "env", "Perwakilan JP Des")
    excel_template = os.path.join(os.getcwd(), "env", "BA_Perwakilan_Template.xlsx")
    output_folder = os.path.join(os.getcwd(), "Output")

    # Menjalankan fungsi utama
    process_all_csv_in_folder(input_folder, excel_template, output_folder)
