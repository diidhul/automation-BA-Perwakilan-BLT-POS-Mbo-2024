import os
from openpyxl import load_workbook
from copy import copy  # Untuk menyalin atribut seperti font, border, alignment


# Fungsi untuk menambahkan footer ke file Excel yang sudah di-generate
def add_footer_to_excel(excel_file, footer_template, output_file):
    try:
        # Memuat file Excel yang sudah di-generate
        workbook = load_workbook(excel_file)
        sheet = workbook.active  # Menggunakan sheet pertama

        # Menemukan baris terakhir yang berisi data
        last_row = sheet.max_row
        print(f"Baris terakhir file '{excel_file}': {last_row}")  # Debug baris terakhir

        # Memuat template footer
        footer_workbook = load_workbook(footer_template)
        footer_sheet = (
            footer_workbook.active
        )  # Menggunakan sheet pertama dari footer template

        # Menyalin footer dari template ke file yang sudah di-generate
        for row_index, row in enumerate(
            footer_sheet.iter_rows(
                min_row=1,
                max_row=footer_sheet.max_row,
                min_col=1,
                max_col=footer_sheet.max_column,
            ),
            start=1,
        ):
            for col, cell in enumerate(row, start=1):
                target_cell = sheet.cell(row=last_row + row_index, column=col)
                target_cell.value = cell.value

                # Menyalin format seperti border, font, alignment sebagai salinan baru
                if cell.border:
                    target_cell.border = copy(cell.border)
                if cell.font:
                    target_cell.font = copy(cell.font)
                if cell.alignment:
                    target_cell.alignment = copy(cell.alignment)

        # Menyalin merge range dari template footer
        for merge_range in footer_sheet.merged_cells.ranges:
            # Offset baris merge range ke file target
            start_row = merge_range.min_row + last_row
            end_row = merge_range.max_row + last_row
            start_col = merge_range.min_col
            end_col = merge_range.max_col
            sheet.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col,
            )

        # Menyimpan file Excel dengan footer yang telah ditambahkan
        workbook.save(output_file)
        return f"Footer berhasil ditambahkan di file: {output_file}"

    except Exception as e:
        return f"Terjadi kesalahan: {e}"


# Fungsi utama untuk menambahkan footer ke semua file Excel di folder output
def add_footer_to_all_excel_files(input_folder, footer_template, output_folder):
    try:
        # Pastikan folder output utama ada
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Iterasi setiap file Excel di folder input
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx"):
                excel_file = os.path.join(input_folder, file_name)
                base_name = os.path.splitext(file_name)[0]  # Nama file tanpa ekstensi

                # Path output file baru dengan footer
                output_file = os.path.join(
                    output_folder, f"{base_name}_with_footer.xlsx"
                )

                # Jalankan penambahan footer
                result = add_footer_to_excel(excel_file, footer_template, output_file)
                print(result)

        print("\nSemua footer udah ada di excell.")
        print("\nSekarang kamu udah bisa leha-lehaa Dini sayaaaaang.")

    except Exception as e:
        print(f"Terjadi kesalahan: {e}")


if __name__ == "__main__":
    # Folder input Excel yang sudah di-generate dan template footer
    input_folder = os.path.join(os.getcwd(), "Output")
    # Folder dengan file Excel yang sudah di-generate
    footer_template = os.path.join(os.getcwd(), "env", "Footer_Template.xlsx")
    # Template footer berada di folder env
    output_folder = os.path.join(os.getcwd(), "Output_With_Footer")
    # Folder output untuk file dengan footer

    # Menjalankan fungsi utama
    add_footer_to_all_excel_files(input_folder, footer_template, output_folder)
